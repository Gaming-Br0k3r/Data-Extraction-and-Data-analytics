import pandas as pd
import requests
import re
import nltk
nltk.download('punkt')
import os
import openpyxl
import shutil
from bs4 import BeautifulSoup
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords

df = pd.read_excel('Input.xlsx')

def File_name(name):
    file = open(name, "w+", encoding='utf-8')
    return file
def count_personal_pronouns(file_path):
    
    pronoun_pattern = r'\b(?:I|we|my|ours|us)\b'

    
    with open(file_path, 'r', encoding='utf-8') as file:
        text_content = file.read()

    matches = re.findall(pronoun_pattern, text_content, flags=re.IGNORECASE)

    pronoun_counts = {}
    for pronoun in set(matches):
        pronoun_counts[pronoun.lower()] = matches.count(pronoun)

    return sum(pronoun_counts.values())
def syllable_count_per_word(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        text = file.read()

    words = re.findall(r'\b\w+\b', text)
    def count_syllables(word):
        word = re.sub(r'(es|ed)$', '', word, flags=re.IGNORECASE)
        vowels = "aeiouAEIOU"
        syllable_count = sum(1 for char in word if char in vowels)
        if syllable_count > 1 and word[-1].lower() == 'e':
            syllable_count -= 1
        return max(1, syllable_count)
    syllable_counts = {word: count_syllables(word) for word in words}
    complex_word_count = sum(1 for syllables in syllable_counts.values() if syllables > 2)
    return complex_word_count,len(syllable_counts) 
def sentiment_analysis(file_path, stop_words_file):
    stop_words_folder = 'StopWords'
    stop_words = set()
    for stop_words_file in os.listdir(stop_words_folder):
        with open(os.path.join(stop_words_folder, stop_words_file), 'r') as file:
            stop_words.update(file.read().split())


    positive_words = set(open('MasterDictionary/positive-words.txt').read().split())
    negative_words = set(open('MasterDictionary/negative-words.txt').read().split())

    with open(file_path, 'r', encoding='utf-8') as text_file:
        text = text_file.read()

    tokens = word_tokenize(text)

    clean_tokens = [word.lower() for word in tokens if word.isalnum() and word.lower() not in stop_words]

    positive_score = sum(1 for word in clean_tokens if word in positive_words)
    
    negative_score = sum(1 for word in clean_tokens if word in negative_words)
    
    polarity_score = (positive_score - negative_score) / ((positive_score + negative_score) + 0.000001)
    
    subjectivity_score = (positive_score + negative_score) / (len(clean_tokens) + 0.000001)
    
    return subjectivity_score,polarity_score,negative_score,positive_score
def calculate_readability_from_file(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            text = file.read()

        cleaned_text = ' '.join(text.split())

        words = re.findall(r'\b\w+\b', cleaned_text)
        num_words = len(words)
        
        total_chars = sum(len(word) for word in words)

        if num_words==0:
            average_word_length = total_chars / 1
        else:
            average_word_length = total_chars / num_words
        
        sentences = re.split(r'[.!?]', cleaned_text)
        num_sentences = len(sentences)

        complex_words = [word for word in words if len(word) > 6]
        num_complex_words = len(complex_words)

        
        average_sentence_length = num_words / num_sentences
        
        percentage_complex_words = (num_complex_words / num_words) * 100
        
        fox_index = 0.4 * (average_sentence_length + percentage_complex_words)
        
        avg_no_words_per_sentences = average_sentence_length
        
        return avg_no_words_per_sentences,fox_index,average_word_length,num_words,percentage_complex_words

    except FileNotFoundError:
        print("File not found. Please provide a valid file path.")
    except Exception as e:
        print(f"An error occurred: {e}")
next_row=2
for index, row in df.iterrows():
    col1 = row['URL_ID']
    col2 = row['URL']

    if isinstance(col1, float) and col1.is_integer():
        n = int(col1)
        id = str(n)
    else:
        id = str(col1)

    id += ".txt"

    test = File_name(id)
    print(f"Scraping URL: {id}")

    try:
        r = requests.get(col2)
        r.raise_for_status()
        x = r.content
        soup = BeautifulSoup(x, "html.parser")
        card = soup.find("div", attrs={"class", "td-post-content tagdiv-type"})

        if card:
            for full in para:
                test.write(str(full.get_text()))
            test.close()
            para = card.find_all("p")
            
            x=id
            y=['StopWords_Auditor.txt', 'StopWords_Currencies.txt', 'StopWords_DatesandNumbers.txt', 'StopWords_Generic.txt', 'StopWords_GenericLong.txt', 'StopWords_Geographic.txt', 'StopWords_Names.txt']
            cw,sc=syllable_count_per_word(x)
            ss,pls,ns,ps=sentiment_analysis(x,y)
            anwpl,ar,awl,wc,pcw=calculate_readability_from_file(x)
            pn=count_personal_pronouns(x)
            print("Now Appending to xcel sheet:")
            folder_path = 'TextFiles'
            os.makedirs(folder_path, exist_ok=True)
            destination_file = os.path.join(folder_path, id)
            os.rename(x, destination_file) 
            workbook=openpyxl.load_workbook('Output Data Structure.xlsx')
            sheet=workbook.active
            sheet.cell(row=next_row, column=3, value=ps)
            sheet.cell(row=next_row, column=4, value=ns)
            sheet.cell(row=next_row, column=5, value=pls)
            sheet.cell(row=next_row, column=6, value=ss)
            sheet.cell(row=next_row, column=7, value=anwpl)
            sheet.cell(row=next_row, column=8, value=pcw)
            sheet.cell(row=next_row, column=9, value=ar)
            sheet.cell(row=next_row, column=10, value=anwpl)
            sheet.cell(row=next_row, column=11, value=cw)
            sheet.cell(row=next_row, column=12, value=wc)
            sheet.cell(row=next_row, column=13, value=sc)
            sheet.cell(row=next_row, column=14, value=pn)
            sheet.cell(row=next_row, column=15, value=awl)
            next_row+= 1
            workbook.save('Output Data Structure.xlsx')

        else:
            print(f"Couldn't find the specified div for {col1}")
            next_row += 1
    except requests.exceptions.RequestException as e:
        print(f"Error scraping URL {col1}: {e}")
        next_row += 1